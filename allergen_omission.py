import openpyxl, re, json, collections

wb = openpyxl.load_workbook('/home/claude/paired_retailer.xlsx', data_only=True)
ws = wb['Ingredients']
hdr = [ (c.value or '') for c in ws[1] ]
def col(sub):
    for i,h in enumerate(hdr):
        if isinstance(h,str) and h.strip().lower().startswith(sub): return i
    return None
iB, iN, iOFFdb = col('barcode'), col('name'), col('off database')
iRaw = col('raw_ingredients')
iApi = col('open_food_facts_api_ingredients')
iMatch = col('match_percentage')

rows=[]
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or r[iB] in (None,''): continue
    rows.append(r)

BIG9 = {
 'milk':      ['milk','cream','butter','whey','casein','lactose','cheese','yogurt','ghee','curd'],
 'egg':       ['egg','albumen','ovalbumin','mayonnaise','meringue'],
 'fish':      ['anchov','tuna','salmon',' cod','tilapia','sardine','fish'],
 'shellfish': ['shrimp','prawn','crab','lobster','crayfish','oyster','clam','mussel','scallop','squid','octopus','krill'],
 'tree nuts': ['almond','cashew','walnut','pecan','pistachio','hazelnut','macadamia','brazil nut','filbert'],
 'peanuts':   ['peanut','groundnut','arachis'],
 'wheat':     ['wheat','semolina','durum','farina','spelt','gluten','couscous','bulgur'],
 'soy':       ['soy','soya','soybean','tofu','edamame','miso','tempeh'],
 'sesame':    ['sesame','tahini','benne'],
}
NEG = ['peanut butter','shea butter','cocoa butter','apple butter','nut butter','almond butter',
       'sunflower butter','soy butter','butternut','coconut','butterfly','cream of tartar']

def tags(text):
    t=(text or '').lower(); m=t
    for n in NEG: m=m.replace(n,' ')
    if 'buttermilk' in t: m+=' buttermilk milk '
    out=set()
    for a,syns in BIG9.items():
        for s in syns:
            if a=='soy' and s=='soy':
                if re.search(r'\bsoy\b',m) or 'soybean' in m or 'soya' in m: out.add(a); break
            elif s in m: out.add(a); break
    return out

def norm(x):
    if x in (None,''): return ''
    s=str(x)
    try:
        v=json.loads(s)
        if isinstance(v,list): return ' ; '.join(str(i) for i in v)
    except Exception: pass
    return s

paired=[]; skipped=collections.Counter()
for r in rows:
    raw, api = norm(r[iRaw]), norm(r[iApi])
    offdb = str(r[iOFFdb]).strip().lower() if r[iOFFdb] else ''
    if not raw: skipped['no own transcription']+=1; continue
    if offdb != 'yes': skipped['not in OFF']+=1; continue
    if not api: skipped['in OFF but API list empty']+=1; continue
    paired.append((str(r[iB]), r[iN], raw, api, r[iMatch] if iMatch else None))

print(f"rows in tab: {len(rows)}")
print(f"PAIRED (own transcription AND OFF API list both present): {len(paired)}")
print("skipped:", dict(skipped)); print()

true_pairs=0; omitted=0; extra=0
per=collections.defaultdict(lambda: collections.Counter())
prod_with_omission=[]
for bc,name,raw,api,mp in paired:
    A_own, A_off = tags(raw), tags(api)
    true_pairs += len(A_own)
    miss = A_own - A_off; add = A_off - A_own
    omitted += len(miss); extra += len(add)
    for a in A_own: per[a]['true']+=1
    for a in miss:  per[a]['omitted']+=1
    if miss: prod_with_omission.append((bc,name,sorted(miss),mp))

print("=== DATABASE-PATH ALLERGEN OMISSION (paired, same barcode) ===")
print(f"true (product, Big-9 allergen) pairs from own transcription : {true_pairs}")
print(f"omitted by the public record                                : {omitted}  = {100*omitted/true_pairs:.1f}%")
print(f"products with >=1 omitted allergen                          : {len(prod_with_omission)}/{len(paired)} = {100*len(prod_with_omission)/len(paired):.1f}%")
print(f"reverse check (in OFF, not in own transcription)            : {extra}")
print()
print("per-allergen  true / omitted")
for a in sorted(per, key=lambda x:-per[x]['true']):
    print(f"  {a:<10} {per[a]['true']:>3} / {per[a]['omitted']:>3}")
print()
print("products with omissions (up to 15):")
for bc,name,miss,mp in prod_with_omission[:15]:
    print(f"  {bc}  {str(name)[:44]:<44} dropped: {', '.join(miss)}")

print()
print("=== SECOND PASS: non-Big-9 dietary-relevant tokens ===")
DIET = {
 'chicken':['chicken'], 'beef':['beef'], 'pork':['pork','bacon','ham '],
 'gelatin':['gelatin','gelatine'], 'lard':['lard','tallow'],
 'shellfish-derived':['carmine','cochineal'],
}
tp=0; om=0; perd=collections.defaultdict(lambda: collections.Counter()); prods=[]
for bc,name,raw,api,mp in paired:
    def dtags(t):
        t=(t or '').lower(); o=set()
        for k,s in DIET.items():
            if any(x in t for x in s): o.add(k)
        return o
    A,B=dtags(raw),dtags(api); tp+=len(A); miss=A-B; om+=len(miss)
    for a in A: perd[a]['true']+=1
    for a in miss: perd[a]['omitted']+=1
    if miss: prods.append((bc,name,sorted(miss)))
print(f"true (product, token) pairs : {tp}")
print(f"omitted by public record    : {om} = {100*om/tp:.1f}%" if tp else "n/a")
for a in sorted(perd,key=lambda x:-perd[x]['true']):
    print(f"  {a:<18} {perd[a]['true']:>3} / {perd[a]['omitted']:>3}")
for bc,name,miss in prods[:10]:
    print(f"  {bc}  {str(name)[:44]:<44} dropped: {', '.join(miss)}")

print()
print("=== ingredient-level overlap on the SAME paired set ===")
mps=[float(m) for _,_,_,_,m in paired if isinstance(m,(int,float))]
if mps:
    mps.sort()
    print(f"n={len(mps)} mean={sum(mps)/len(mps):.1f}% median={mps[len(mps)//2]:.1f}% "
          f"<50%: {sum(1 for x in mps if x<50)} | <25%: {sum(1 for x in mps if x<25)} | ==100%: {sum(1 for x in mps if x>=100)}")
