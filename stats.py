import math, json, csv, re, collections
import openpyxl
from scipy import stats as st

def wilson(k, n, z=1.959963985):
    if n == 0: return (float('nan'),)*3
    p = k/n; d = 1 + z*z/n
    c = (p + z*z/(2*n))/d
    h = z*math.sqrt(p*(1-p)/n + z*z/(4*n*n))/d
    return p, max(0, c-h), min(1, c+h)

def fmt(k, n, label):
    p, lo, hi = wilson(k, n)
    print(f"{label:<46} {k}/{n} = {100*p:5.1f}%   95% CI [{100*lo:4.1f}, {100*hi:4.1f}]")
    return p, lo, hi

print("="*78); print("DATABASE PATH (paired, n=51 products)"); print("="*78)

wb = openpyxl.load_workbook('paired_retailer.xlsx', data_only=True)
ws = wb['Ingredients']; hdr=[(c.value or '') for c in ws[1]]
def col(sub):
    for i,h in enumerate(hdr):
        if isinstance(h,str) and h.strip().lower().startswith(sub): return i
iB,iOFF,iRaw,iApi,iMatch = col('barcode'),col('off database'),col('raw_ingredients'),col('open_food_facts_api_ingredients'),col('match_percentage')

BIG9={'milk':['milk','cream','butter','whey','casein','lactose','cheese','yogurt','ghee','curd'],
 'egg':['egg','albumen','ovalbumin','mayonnaise','meringue'],
 'fish':['anchov','tuna','salmon','tilapia','sardine','fish'],
 'shellfish':['shrimp','prawn','crab','lobster','crayfish','oyster','clam','mussel','scallop','squid','krill'],
 'tree nuts':['almond','cashew','walnut','pecan','pistachio','hazelnut','macadamia','brazil nut','filbert'],
 'peanuts':['peanut','groundnut','arachis'],
 'wheat':['wheat','semolina','durum','farina','spelt','gluten','couscous','bulgur'],
 'soy':['soy','soya','soybean','tofu','edamame','miso','tempeh'],
 'sesame':['sesame','tahini','benne']}
NEG=['peanut butter','shea butter','cocoa butter','apple butter','nut butter','almond butter',
     'sunflower butter','soy butter','butternut','coconut','cream of tartar']
def tags(x):
    t=(x or '').lower(); m=t
    for n in NEG: m=m.replace(n,' ')
    if 'buttermilk' in t: m+=' buttermilk milk '
    o=set()
    for a,syn in BIG9.items():
        for s in syn:
            if a=='soy' and s=='soy':
                if re.search(r'\bsoy\b',m) or 'soybean' in m or 'soya' in m: o.add(a); break
            elif s in m: o.add(a); break
    return o
def norm(x):
    if x in (None,''): return ''
    s=str(x)
    try:
        v=json.loads(s)
        if isinstance(v,list): return ' ; '.join(str(i) for i in v)
    except Exception: pass
    return s

paired=[]
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or r[iB] in (None,''): continue
    raw,api = norm(r[iRaw]), norm(r[iApi])
    if not raw or str(r[iOFF]).strip().lower()!='yes' or not api: continue
    paired.append((raw, api, r[iMatch]))

mp=[float(m) for _,_,m in paired if isinstance(m,(int,float))]
n=len(mp); mean=sum(mp)/n; sd=math.sqrt(sum((x-mean)**2 for x in mp)/(n-1))
tcrit=st.t.ppf(0.975, n-1); half=tcrit*sd/math.sqrt(n)
print(f"ingredient overlap: n={n}  mean={mean:.1f}%  SD={sd:.1f}  95% CI [{mean-half:.1f}, {mean+half:.1f}]")
print(f"                    median={sorted(mp)[n//2]:.1f}%  IQR [{sorted(mp)[n//4]:.1f}, {sorted(mp)[3*n//4]:.1f}]")
disagree = 100-mean
print(f"ingredient-level DISAGREEMENT = {disagree:.1f}%")

tp=om=extra=0
for raw,api,_ in paired:
    A,B=tags(raw),tags(api); tp+=len(A); om+=len(A-B); extra+=len(B-A)
p_om,lo_om,hi_om = fmt(om, tp, "allergen omission (pairs)")
print(f"reverse direction (in record, not ours)        {extra}")
# is the allergen-level rate lower than the ingredient-level rate?
print(f"\nOne-sided binomial test, H0: allergen omission >= ingredient disagreement ({disagree/100:.3f})")
pv = st.binomtest(om, tp, disagree/100, alternative='less').pvalue
print(f"  p = {pv:.2e}   -> {'reject H0' if pv<0.05 else 'cannot reject'}")
print(f"  ratio of point estimates = {disagree/(100*p_om):.1f}x  (CI on ratio is wide; see hi bound {100*hi_om:.1f}%)")

print(); print("="*78); print("CAMERA PATH"); print("="*78)
rows=list(csv.DictReader(open('ds/labels200.csv')))
def toks(raw):
    try:
        v=json.loads(raw)
        if isinstance(v,list): return [str(x) for x in v]
    except Exception: pass
    return [raw] if raw else []
def hit(tok,a):
    t=tok.lower()
    for x in NEG: t=t.replace(x,' ')
    if 'buttermilk' in tok.lower(): t+=' milk '
    for s in BIG9[a]:
        if a=='soy' and s=='soy':
            if re.search(r'\bsoy\b',t) or 'soybean' in t or 'soya' in t: return True
        elif s in t: return True
    return False
kd=collections.Counter(); tot=0; per=collections.defaultdict(lambda: collections.Counter())
for r in rows:
    tl=toks(r['ingredients'])
    for a in BIG9:
        k=sum(1 for tk in tl if hit(tk,a))
        if k:
            tot+=1; kd[min(k,4)]+=1; per[a]['n']+=1
            if k==1: per[a]['k1']+=1
fmt(kd[1], tot, "verdicts resting on a single token (k=1)")
print()
print("per allergen, k=1 share with 95% CI:")
for a in sorted(per,key=lambda x:-per[x]['n']):
    fmt(per[a]['k1'], per[a]['n'], "  "+a)

print()
n_cam=262; obs=round(0.408*n_cam); pred=0.308
p,lo,hi = fmt(obs, n_cam, "camera-path allergen FNR (observed)")
pv2 = st.binomtest(obs, n_cam, pred, alternative='greater').pvalue
print(f"vs independence model {100*pred:.1f}%: one-sided binomial p = {pv2:.2e}")

print(); print("="*78); print("AGGREGATE RISK, with uncertainty propagated"); print("="*78)
cov_k, cov_n = 57, 67
c,clo,chi = wilson(cov_k,cov_n)
print(f"coverage c = {100*c:.1f}%  95% CI [{100*clo:.1f}, {100*chi:.1f}]")
def R(c,rdb,rcam): return c*rdb+(1-c)*rcam
base = R(c,p_om,p)
lo_R = min(R(cc,rr,rc) for cc in (clo,c,chi) for rr in (lo_om,p_om,hi_om) for rc in (lo,p,hi))
hi_R = max(R(cc,rr,rc) for cc in (clo,c,chi) for rr in (lo_om,p_om,hi_om) for rc in (lo,p,hi))
print(f"aggregate risk R = {100*base:.1f}%   plausible range [{100*lo_R:.1f}, {100*hi_R:.1f}] over the CI box")
print(f"camera share of R at point estimate = {100*(1-c)*p/base:.0f}%")
share_lo = min((1-cc)*rc/R(cc,rr,rc) for cc in (clo,c,chi) for rr in (lo_om,p_om,hi_om) for rc in (lo,p,hi))
share_hi = max((1-cc)*rc/R(cc,rr,rc) for cc in (clo,c,chi) for rr in (lo_om,p_om,hi_om) for rc in (lo,p,hi))
print(f"camera share range                  = [{100*share_lo:.0f}, {100*share_hi:.0f}]%")
